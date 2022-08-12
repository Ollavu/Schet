from django.contrib import admin
admin.site.site_header = 'Giorno Giovanna'
from import_export.admin import ImportExportModelAdmin
from import_export import resources
from import_export.admin import ExportActionMixin
from django.contrib.auth.models import PermissionsMixin
from .models import *
from top.views import GetOpenIdByStr, GetObjectt
from django.views.decorators.cache import cache_page
from django.contrib.admin import SimpleListFilter
from functools import lru_cache, wraps
from datetime import datetime, timedelta
from openpyxl import Workbook
import getpass
from django.contrib import messages
from top.functions import GetPartOfTime


class OpeningAdmin(admin.ModelAdmin):
	list_display = ('location','workerid','date','time')
	list_display_links = ('location','workerid','date','time',)
	search_fields = ('date', 'time')
	list_filter = ('openid','workerid','objectid', 'date','time','location','distance')
	fields = ('openid','workerid','objectid', 'date','time','location','distance')
	readonly_fields = fields

class WorkerAdmin(admin.ModelAdmin):
	list_display = ('fullname','user','post','tabel')
	list_display_links = ('fullname','tabel',)
	search_fields = ('post', 'tabel')
	save_on_top = True

class ObjectAdmin(admin.ModelAdmin):
	list_display = ('name','objecttype','radius')
	list_display_links = ('name','objecttype',)
	search_fields = ('objecttype', 'radius')
	save_on_top = True


def timed_lru_cache(seconds: int, maxsize: int = 128):
		def wrapper_cache(func):
				func = lru_cache(maxsize=maxsize)(func)
				func.lifetime = timedelta(seconds=seconds)
				func.expiration = datetime.utcnow() + func.lifetime

				@wraps(func)
				def wrapped_func(*args, **kwargs):
						if datetime.utcnow() >= func.expiration:
								func.cache_clear()
								func.expiration = datetime.utcnow() + func.lifetime

						return func(*args, **kwargs)

				return wrapped_func

		return wrapper_cache

#ЗАКРЫТЫЕ РАБОТЯГИ
class ClosingAdmin(admin.ModelAdmin):

	list_display = ('openid','distance','date','time','statusid')
	list_display_links = ('openid','distance','date','time',)

	list_filter = ('date', 'statusid')
	fields = ('noteid','openid','time', 'distance','location','photo_1','photo_2','photo_3','statusid','date')
	readonly_fields = ('noteid','openid','time','date','location','distance','photo_1','photo_2','photo_3')

	actions = ('synchronization','full_time_report',)

	def synchronization(self, request, queryset):
		for note in queryset:
			if str(note.statusid) == 'Принята':
				export = DataToExport()

				open_id = GetOpenIdByStr(str(note))

				export.worker_fullname = str(open_id.workerid)
				export.object_name = str(open_id.objectid)
				export.opendate = open_id.date
				export.closedate = note.date
				export.spended_time = datetime.combine(note.date,note.time) - datetime.combine(open_id.date,open_id.time)

				#проверка на уникальную запись
				a= DataToExport.objects.filter(worker_fullname = export.worker_fullname, object_name=export.object_name,  opendate=export.opendate,  closedate=export.closedate, spended_time=export.spended_time)

				if str(a) == str(DataToExport.objects.none()):
					export.save()

				del export
				del a
		self.message_user(request, 'Закрытые заявки синхронизированы')

	def full_time_report(self, request, queryset):

		unique_objects = []

		for note in queryset:
			if str(note.statusid) == 'Принята':

				export = TotalSpendedTimeOnObject()
				open_id = GetOpenIdByStr(str(note))

				if str(open_id.objectid) in unique_objects:
					del export
					del open_id
					continue
				else:
					unique_objects.append(str(open_id.objectid))
					fulltime = []
					for noted in queryset:
						open_id2 = GetOpenIdByStr(str(noted))

						if str(open_id2.objectid) == str(open_id.objectid):
							fulltime.append(datetime.combine(noted.date,noted.time)-datetime.combine(open_id2.date,open_id2.time))


				export.worker_fullname = str(open_id.workerid)
				export.object_name = str(open_id.objectid)
				export.spended_time = fulltime[0]

				#СУММИРОВАНИЕ ВСЕГО ЗАТРАЧЕННОГО ВРЕМЕНИ НА ОБЪЕКТ
				for x in range(1,len(fulltime)):
					export.spended_time+=fulltime[x]

				export.spended_time = str(export.spended_time)

				#проверка на уникальную запись
				a = TotalSpendedTimeOnObject.objects.filter(worker_fullname = export.worker_fullname, object_name = export.object_name, spended_time = str(export.spended_time))

				#ЕСЛИ НЕМА ТО ПИШЕМ
				#print(str(TotalSpendedTimeOnObject.objects.none()))
				if str(a) == str(TotalSpendedTimeOnObject.objects.none()):
					export.save()
					pass

				del export
				del a
				del fulltime

		self.message_user(request, 'Закрытые заявки синхронизированы')


	@timed_lru_cache(120)
	def ff(self, obj):
		return obj.Opening.workerid


	@timed_lru_cache(120)
	def workername(self, rec):
		return str(GetOpenIdByStr(str(rec.openid)).workerid)
	workername.short_description = 'ФИО'


	@timed_lru_cache(120)
	def objectname(self, rec):
		return str(GetOpenIdByStr(str(rec.openid)).objectid)
	objectname.short_description = 'Объект'


	@timed_lru_cache(120)
	def opentime(self, rec):
		return str(GetOpenIdByStr(str(rec.openid)).time)
	opentime.short_description = 'Время открытия'


	@timed_lru_cache(120)
	def opendate(self, rec):
		return str(GetOpenIdByStr(str(rec.openid)).date)
	opendate.short_description = 'Дата открытия'

	def discount(self, request, queryset):
		for rec in queryset:
			print(GetOpenIdByStr(str(rec.openid)).workerid)
		self.message_user(request, 'Dействие выполнено')
	discount.short_description = 'Уменьшить цену вдвое'

class AdminClosedToExport(ExportActionMixin, admin.ModelAdmin ):

	list_display = ('worker_fullname','object_name','opendate','closedate','spended_time')
	list_display_links = ('worker_fullname','object_name','opendate','closedate',)
	list_filter = ('worker_fullname','object_name','opendate','closedate')

	actions = ('export_requests_to_xlsx',)


	def export_requests_to_xlsx(self, request, queryset):

		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'lol'

		columns = [
		'ФИО',
		'Объект',
		'Заявка открыта',
		'Заявка закрыта',
		'Затраченное время',
		]

		row_num = 1

		for col_num, column_title in enumerate(columns, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = column_title

		for note in queryset:
			row_num += 1
			
			row = [
			note.worker_fullname,
			note.object_name,
			note.opendate,
			note.closedate,
			note.spended_time,
			]

			for col_num, cell_value in enumerate(row, 1):
				cell = worksheet.cell(row=row_num, column=col_num)
				cell.value = cell_value

		workbook.save('C:/Users/' + getpass.getuser()+ '/Desktop/' + str(datetime.now().year)+"Полный вывод.xlsx")
		self.message_user(request, 'Записано в Excel')

	export_requests_to_xlsx.short_description = 'Записать в Excel'

class AdminFulltimeObjectsToExport(admin.ModelAdmin ):

	list_display = ('worker_fullname','object_name','spended_time')
	list_display_links = ('worker_fullname','object_name',)
	list_filter = ('worker_fullname','object_name','spended_time')

	actions = ('export_fulltime_to_xlsx,')

	def export_fulltime_to_xlsx(self, request, queryset):

		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Все часы по объекту'

		columns = [
		'ФИО',
		'Объект',
		'Полное время',
		]

		row_num = 1

		for col_num, column_title in enumerate(columns, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = column_title

		for note in queryset:
			row_num += 1

			row = [
			note.worker_fullname,
			note.object_name,
			note.spended_time,
			]

			for col_num, cell_value in enumerate(row, 1):
				cell = worksheet.cell(row=row_num, column=col_num)
				cell.value = cell_value

		workbook.save('C:/Users/' + getpass.getuser()+ '/Desktop/' + str(datetime.now().year)+"Вывод со всем временем.xlsx")
		self.message_user(request, 'Записано в Excel')

	export_fulltime_to_xlsx.short_description = 'Полная выборка по объетам'

admin.site.register(Opening, OpeningAdmin)
admin.site.register(Worker, WorkerAdmin)
admin.site.register(Object, ObjectAdmin)
admin.site.register(Closing, ClosingAdmin)
admin.site.register(ClosingStatus)
admin.site.register(DataToExport, AdminClosedToExport)
admin.site.register(TotalSpendedTimeOnObject, AdminFulltimeObjectsToExport)





